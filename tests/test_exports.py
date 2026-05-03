"""Tests for PDF/Excel export services and endpoints.

These tests stand alone: they don't rely on ``app.create_app`` (the project
factory has unresolved imports in this branch) and instead build a minimal
Flask app that wires up the SQLAlchemy ``db`` instance and registers the
exports namespace. That keeps the export feature testable independently of
upstream wiring fixes.
"""

import os
import sys
import types
from datetime import date
from io import BytesIO

import pytest

# ---------------------------------------------------------------------------
# Stub out broken imports so that ``app.models``/``app.api.exports`` load.
# The real `app.utils.cache` module is missing in this worktree but is
# imported by ``app.utils.__init__``; we provide a minimal stand-in.
# ---------------------------------------------------------------------------
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

if "app.utils.cache" not in sys.modules:
    cache_stub = types.ModuleType("app.utils.cache")

    class _CacheManager:
        redis_client = None

        def init_app(self, app):
            return None

    def _cached(*args, **kwargs):  # decorator factory passthrough
        def decorator(fn):
            return fn
        return decorator

    cache_stub.cache_manager = _CacheManager()
    cache_stub.cached = _cached
    sys.modules["app.utils.cache"] = cache_stub


# Provide a stub for ``app.models.prospect``: the module is imported by
# ``app.models.__init__`` but doesn't exist in this branch.
if "app.models.prospect" not in sys.modules:
    prospect_stub = types.ModuleType("app.models.prospect")

    class _Stub:
        pass

    prospect_stub.ProspectLeague = _Stub
    prospect_stub.MinorLeagueTeam = _Stub
    prospect_stub.Prospect = _Stub
    prospect_stub.ProspectStat = _Stub
    sys.modules["app.models.prospect"] = prospect_stub


# Same for ``app.models.api_key``.
if "app.models.api_key" not in sys.modules:
    api_key_stub = types.ModuleType("app.models.api_key")

    class _ApiKeyStub:
        pass

    api_key_stub.ApiKey = _ApiKeyStub
    sys.modules["app.models.api_key"] = api_key_stub


# ``app.utils.security`` is referenced by other api modules but missing.
if "app.utils.security" not in sys.modules:
    security_stub = types.ModuleType("app.utils.security")

    def _require_api_key(fn):
        return fn

    security_stub.require_api_key = _require_api_key
    sys.modules["app.utils.security"] = security_stub


# ``app.api.keys`` and ``app.api.prospects`` are imported but missing.
for missing in ("app.api.keys", "app.api.prospects"):
    if missing not in sys.modules:
        sys.modules[missing] = types.ModuleType(missing)


# Now safe to import the rest
from flask import Flask  # noqa: E402
from flask_login import LoginManager  # noqa: E402

from app import db  # noqa: E402
from app.models import (  # noqa: E402
    AthleteProfile,
    AthleteSkill,
    AthleteStat,
    Position,
    Sport,
    User,
)
from app.services.excel_export import (  # noqa: E402
    athlete_profile_xlsx,
    rankings_xlsx,
    search_results_xlsx,
)
from app.services.pdf_export import (  # noqa: E402
    athlete_profile_pdf,
    rankings_pdf,
    search_results_pdf,
)

PDF_MAGIC = b"%PDF"
XLSX_MAGIC = b"PK\x03\x04"  # zip-based xlsx


# ---------------------------------------------------------------------------
# App / fixture wiring
# ---------------------------------------------------------------------------

@pytest.fixture
def app():
    flask_app = Flask(__name__)
    flask_app.config.update(
        TESTING=True,
        SECRET_KEY="export-test-key",
        SQLALCHEMY_DATABASE_URI="sqlite:///:memory:",
        SQLALCHEMY_TRACK_MODIFICATIONS=False,
        SESSION_COOKIE_SECURE=False,
    )
    db.init_app(flask_app)
    lm = LoginManager()
    lm.init_app(flask_app)

    @lm.user_loader
    def _load_user(user_id):
        return User.query.get(user_id)

    # Register the exports namespace on its own Flask-RESTX Api so the
    # endpoints can be exercised without importing the project's broken
    # api/__init__.py.
    from flask_restx import Api

    from app.api import exports as exports_module

    blueprint = flask_app.blueprints  # avoid duplicate registration on reuse
    if "exports_api" not in blueprint:
        bp_api = Api(flask_app, doc=False)
        bp_api.add_namespace(exports_module.ns, path="/api/exports")

    with flask_app.app_context():
        db.create_all()
        _seed(flask_app)
        yield flask_app
        db.session.remove()
        db.drop_all()


@pytest.fixture
def client(app):
    return app.test_client()


@pytest.fixture
def auth_headers(app):
    """Build a bearer-token header by inserting a synthetic OAuth account."""
    from app.models.oauth import UserOAuthAccount

    with app.app_context():
        user = User.query.filter_by(username="exporter").first()
        if not user:
            user = User(
                username="exporter",
                email="exporter@example.com",
                first_name="Ex",
                last_name="Porter",
            )
            user.save()
        if not UserOAuthAccount.query.filter_by(access_token="export-token").first():
            account = UserOAuthAccount(
                user_id=user.user_id,
                provider_name="test",
                provider_user_id="export-1",
                access_token="export-token",
            )
            db.session.add(account)
            db.session.commit()
    return {"Authorization": "Bearer export-token"}


# ---------------------------------------------------------------------------
# Synthetic athlete fixture
# ---------------------------------------------------------------------------

def _seed(app):
    sport = Sport(name="Basketball", code="NBA")
    db.session.add(sport)
    db.session.flush()
    position = Position(sport_id=sport.sport_id, name="Forward", code="SF")
    db.session.add(position)
    db.session.flush()

    user = User(
        username="lebron",
        email="lebron@example.com",
        first_name="LeBron",
        last_name="James",
    )
    user.save()
    athlete = AthleteProfile(
        user_id=user.user_id,
        primary_sport_id=sport.sport_id,
        primary_position_id=position.position_id,
        date_of_birth=date(1984, 12, 30),
        height_cm=206,
        weight_kg=113.4,
        nationality="USA",
        current_team="Los Angeles Lakers",
        jersey_number="23",
        bio="Forward known for elite court vision and scoring.",
        overall_rating=98.5,
    )
    athlete.save()
    db.session.add_all([
        AthleteStat(athlete_id=athlete.athlete_id, name="PointsPerGame",
                    value="27.5", stat_type="season", season="2024"),
        AthleteStat(athlete_id=athlete.athlete_id, name="AssistsPerGame",
                    value="8.1", stat_type="season", season="2024"),
        AthleteSkill(athlete_id=athlete.athlete_id, name="Finishing", level=9),
        AthleteSkill(athlete_id=athlete.athlete_id, name="Vision", level=10),
    ])
    db.session.commit()
    app.config["EXPORT_ATHLETE_ID"] = athlete.athlete_id


# ---------------------------------------------------------------------------
# Service-level tests (no HTTP)
# ---------------------------------------------------------------------------

def test_athlete_profile_pdf_renders(app):
    aid = app.config["EXPORT_ATHLETE_ID"]
    with app.app_context():
        buf = athlete_profile_pdf(aid)
    assert isinstance(buf, BytesIO)
    data = buf.getvalue()
    assert data.startswith(PDF_MAGIC)
    assert len(data) > 1500  # non-trivial document


def test_athlete_profile_xlsx_renders(app):
    aid = app.config["EXPORT_ATHLETE_ID"]
    with app.app_context():
        buf = athlete_profile_xlsx(aid)
    data = buf.getvalue()
    assert data.startswith(XLSX_MAGIC)

    # Parse it back to verify sheet structure.
    from openpyxl import load_workbook
    buf.seek(0)
    wb = load_workbook(buf)
    assert {"Summary", "Stats", "Games", "Skills"}.issubset(set(wb.sheetnames))

    stats_headers = [c.value for c in wb["Stats"][1]]
    assert stats_headers[:4] == ["Season", "Stat", "Value", "Type"]

    skill_headers = [c.value for c in wb["Skills"][1]]
    assert skill_headers[:2] == ["Skill", "Level"]


def test_search_results_pdf_renders(app):
    with app.app_context():
        athletes = AthleteProfile.query.all()
        buf = search_results_pdf(athletes, {"sport": "NBA", "min_age": 30})
    data = buf.getvalue()
    assert data.startswith(PDF_MAGIC)


def test_search_results_xlsx_renders(app):
    with app.app_context():
        athletes = AthleteProfile.query.all()
        buf = search_results_xlsx(athletes, {"sport": "NBA"})
    data = buf.getvalue()
    assert data.startswith(XLSX_MAGIC)
    from openpyxl import load_workbook
    buf.seek(0)
    wb = load_workbook(buf)
    assert {"Filters", "Results"}.issubset(set(wb.sheetnames))
    headers = [c.value for c in wb["Results"][1]]
    assert headers[:3] == ["#", "Athlete ID", "Name"]


def test_rankings_pdf_renders():
    rows = [
        {"name": "LeBron James", "score": 98.5, "team": "LAL"},
        {"name": "Stephen Curry", "score": 94.9, "team": "GSW"},
    ]
    weights = {"Offense": 0.4, "Defense": 0.3, "Durability": 0.2, "Intangibles": 0.1}
    buf = rankings_pdf(rows, sport="nba", weights=weights)
    assert buf.getvalue().startswith(PDF_MAGIC)


def test_rankings_xlsx_renders():
    rows = [{"name": "LeBron James", "score": 98.5, "team": "LAL"}]
    buf = rankings_xlsx(rows, sport="nba", weights={"OverallRating": 1.0})
    data = buf.getvalue()
    assert data.startswith(XLSX_MAGIC)
    from openpyxl import load_workbook
    buf.seek(0)
    wb = load_workbook(buf)
    assert "Rankings" in wb.sheetnames
    assert "Weights" in wb.sheetnames
    headers = [c.value for c in wb["Rankings"][4]]
    assert headers[:4] == ["Rank", "Athlete", "Team", "Score"]


def test_athlete_profile_pdf_missing_id(app):
    with app.app_context(), pytest.raises(ValueError):
        athlete_profile_pdf("does-not-exist")


# ---------------------------------------------------------------------------
# HTTP endpoint tests
# ---------------------------------------------------------------------------

def test_athlete_pdf_endpoint(client, app, auth_headers):
    aid = app.config["EXPORT_ATHLETE_ID"]
    resp = client.get(f"/api/exports/athletes/{aid}.pdf", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.mimetype == "application/pdf"
    assert resp.data.startswith(PDF_MAGIC)
    assert len(resp.data) > 1000
    cd = resp.headers.get("Content-Disposition", "")
    assert "attachment" in cd
    assert ".pdf" in cd


def test_athlete_xlsx_endpoint(client, app, auth_headers):
    aid = app.config["EXPORT_ATHLETE_ID"]
    resp = client.get(f"/api/exports/athletes/{aid}.xlsx", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.mimetype.endswith("spreadsheetml.sheet")
    assert resp.data.startswith(XLSX_MAGIC)
    assert "attachment" in resp.headers.get("Content-Disposition", "")


def test_search_pdf_endpoint(client, auth_headers):
    resp = client.get("/api/exports/search.pdf?sport=NBA", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.mimetype == "application/pdf"
    assert resp.data.startswith(PDF_MAGIC)


def test_search_xlsx_endpoint(client, auth_headers):
    resp = client.get("/api/exports/search.xlsx?sport=NBA", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.data.startswith(XLSX_MAGIC)


def test_rankings_pdf_endpoint(client, auth_headers):
    resp = client.get(
        "/api/exports/rankings.pdf?sport=NBA&preset_id=balanced",
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert resp.data.startswith(PDF_MAGIC)


def test_rankings_xlsx_endpoint(client, auth_headers):
    resp = client.get(
        "/api/exports/rankings.xlsx?sport=NBA&preset_id=balanced",
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert resp.data.startswith(XLSX_MAGIC)


def test_endpoints_require_auth(client):
    aid_resp = client.get("/api/exports/search.pdf")
    assert aid_resp.status_code == 401
