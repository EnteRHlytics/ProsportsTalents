"""Excel (.xlsx) export service.

Mirrors :mod:`app.services.pdf_export` so the agency can hand a sponsor
either format. Multi-sheet workbooks are used wherever a single tab would
flatten useful structure (e.g. an athlete profile gets Summary / Stats /
Games / Skills tabs).
"""

from __future__ import annotations

from collections.abc import Iterable
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="FF374151", end_color="FF374151", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="FFE07A1F", end_color="FFE07A1F", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color="FF0F1F3D", size=16)
LABEL_FONT = Font(name="Calibri", bold=True, color="FF374151", size=10)
BODY_FONT = Font(name="Calibri", size=10)
BORDER = Border(
    left=Side(style="thin", color="FFD1D5DB"),
    right=Side(style="thin", color="FFD1D5DB"),
    top=Side(style="thin", color="FFD1D5DB"),
    bottom=Side(style="thin", color="FFD1D5DB"),
)


def _style_header_row(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER


def _apply_body_borders(ws, start_row: int, end_row: int, ncols: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if cell.font is None or cell.font.size != 11:
                cell.font = BODY_FONT


def _autosize(ws, max_width: int = 40) -> None:
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def _write_table(ws, start_row: int, headers: list, rows: list) -> int:
    """Write a header row then body rows. Returns the row after the table."""
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h)
    _style_header_row(ws, start_row, len(headers))
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            ws.cell(row=start_row + i, column=j, value=val)
    if rows:
        _apply_body_borders(
            ws, start_row + 1, start_row + len(rows), len(headers)
        )
    return start_row + len(rows) + 1


# ---------------------------------------------------------------------------
# Athlete profile data extraction (kept local to avoid pdf import cycle)
# ---------------------------------------------------------------------------

def _basics(athlete) -> dict:
    user = getattr(athlete, "user", None)
    name = user.full_name if user else f"Athlete {getattr(athlete, 'athlete_id', '')}"
    sport = getattr(athlete, "primary_sport", None)
    position = getattr(athlete, "primary_position", None)
    return {
        "athlete_id": getattr(athlete, "athlete_id", None),
        "name": name,
        "sport": getattr(sport, "name", None) or getattr(sport, "code", None) or "",
        "position": getattr(position, "name", None) or getattr(position, "code", None) or "",
        "team": getattr(athlete, "current_team", None) or "Free Agent",
        "jersey": getattr(athlete, "jersey_number", None) or "",
        "dob": getattr(athlete, "date_of_birth", None),
        "age": getattr(athlete, "age", None),
        "height_cm": getattr(athlete, "height_cm", None),
        "weight_kg": float(athlete.weight_kg) if getattr(athlete, "weight_kg", None) is not None else None,
        "nationality": getattr(athlete, "nationality", None) or "",
        "bio": getattr(athlete, "bio", None) or "",
        "rating": float(athlete.overall_rating) if getattr(athlete, "overall_rating", None) is not None else None,
        "email": getattr(user, "email", None) if user else "",
    }


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def athlete_profile_xlsx(athlete_id: str) -> BytesIO:
    """Return a multi-sheet workbook for a single athlete."""
    from app.models import AthleteProfile, AthleteSkill, AthleteStat

    athlete = (
        AthleteProfile.query
        .filter_by(athlete_id=athlete_id, is_deleted=False)
        .first()
    )
    if athlete is None:
        raise ValueError(f"Athlete {athlete_id} not found")

    season_stats = (
        AthleteStat.query
        .filter_by(athlete_id=athlete_id)
        .order_by(AthleteStat.season.desc().nullslast(), AthleteStat.name)
        .all()
    )
    skills = (
        AthleteSkill.query
        .filter_by(athlete_id=athlete_id)
        .order_by(AthleteSkill.level.desc().nullslast(), AthleteSkill.name)
        .all()
    )

    wb = Workbook()

    # ----- Summary -----------------------------------------------------
    summary = wb.active
    summary.title = "Summary"
    info = _basics(athlete)
    summary["A1"] = info["name"]
    summary["A1"].font = TITLE_FONT
    summary.merge_cells("A1:D1")
    summary["A2"] = "Pro Sports Talents | Athlete Profile"
    summary["A2"].font = Font(italic=True, color="FF6B7280")
    summary.merge_cells("A2:D2")

    pairs = [
        ("Athlete ID", info["athlete_id"]),
        ("Sport", info["sport"]),
        ("Position", info["position"]),
        ("Team", info["team"]),
        ("Jersey", info["jersey"]),
        ("Date of Birth", info["dob"]),
        ("Age", info["age"]),
        ("Height (cm)", info["height_cm"]),
        ("Weight (kg)", info["weight_kg"]),
        ("Nationality", info["nationality"]),
        ("Overall Rating", info["rating"]),
        ("Email", info["email"]),
    ]
    for i, (label, value) in enumerate(pairs, start=4):
        a = summary.cell(row=i, column=1, value=label)
        a.font = LABEL_FONT
        a.fill = PatternFill(start_color="FFF3F4F6", end_color="FFF3F4F6", fill_type="solid")
        a.border = BORDER
        b = summary.cell(row=i, column=2, value=value)
        b.font = BODY_FONT
        b.border = BORDER

    bio_row = 4 + len(pairs) + 1
    summary.cell(row=bio_row, column=1, value="Biography").font = LABEL_FONT
    summary.cell(row=bio_row + 1, column=1, value=info["bio"]).alignment = Alignment(wrap_text=True, vertical="top")
    summary.merge_cells(start_row=bio_row + 1, start_column=1, end_row=bio_row + 5, end_column=4)
    summary.row_dimensions[bio_row + 1].height = 70
    _autosize(summary)

    # ----- Stats -------------------------------------------------------
    ws_stats = wb.create_sheet("Stats")
    rows = [
        [s.season or "", s.name or "", s.value if s.value is not None else "", s.stat_type or ""]
        for s in season_stats
    ]
    _write_table(ws_stats, 1, ["Season", "Stat", "Value", "Type"], rows)
    _autosize(ws_stats)

    # ----- Games -------------------------------------------------------
    ws_games = wb.create_sheet("Games")
    games = _recent_games_for(athlete, limit=20)
    grows = []
    for g in games:
        home = _team_label(getattr(g, "home_team", None))
        away = _team_label(getattr(g, "visitor_team", None))
        grows.append([
            g.date.isoformat() if getattr(g, "date", None) else "",
            home,
            g.home_team_score if g.home_team_score is not None else "",
            g.visitor_team_score if g.visitor_team_score is not None else "",
            away,
        ])
    _write_table(ws_games, 1, ["Date", "Home", "Home Score", "Away Score", "Away"], grows)
    _autosize(ws_games)

    # ----- Skills ------------------------------------------------------
    ws_skills = wb.create_sheet("Skills")
    srows = [[sk.name or "", sk.level if sk.level is not None else ""] for sk in skills]
    _write_table(ws_skills, 1, ["Skill", "Level"], srows)
    _autosize(ws_skills)

    return _to_bytes(wb)


def search_results_xlsx(athletes_list: Iterable, filters_summary: dict) -> BytesIO:
    """Search results workbook: ``Filters`` sheet + ``Results`` sheet."""
    wb = Workbook()
    ws_filters = wb.active
    ws_filters.title = "Filters"
    ws_filters["A1"] = "Athlete Search Results"
    ws_filters["A1"].font = TITLE_FONT
    ws_filters.merge_cells("A1:B1")
    ws_filters["A2"] = "Pro Sports Talents | Curated short list"
    ws_filters["A2"].font = Font(italic=True, color="FF6B7280")
    ws_filters.merge_cells("A2:B2")

    rows = []
    if filters_summary:
        for k, v in filters_summary.items():
            if v in (None, "", []):
                continue
            rows.append([str(k).replace("_", " ").title(), str(v)])
    if not rows:
        rows.append(["(none)", "All athletes"])
    _write_table(ws_filters, 4, ["Filter", "Value"], rows)

    athletes_list = list(athletes_list)
    ws_filters.cell(
        row=5 + len(rows) + 1, column=1,
        value=f"Total results: {len(athletes_list)}",
    ).font = LABEL_FONT
    ws_filters.cell(
        row=5 + len(rows) + 2, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = Font(italic=True, color="FF6B7280")
    _autosize(ws_filters)

    ws_results = wb.create_sheet("Results")
    headers = ["#", "Athlete ID", "Name", "Sport", "Position", "Team",
               "Age", "Height (cm)", "Weight (kg)", "Rating"]
    rrows = []
    for i, ath in enumerate(athletes_list, start=1):
        info = _basics(ath)
        rrows.append([
            i,
            info["athlete_id"],
            info["name"],
            info["sport"],
            info["position"],
            info["team"],
            info["age"],
            info["height_cm"],
            info["weight_kg"],
            info["rating"],
        ])
    _write_table(ws_results, 1, headers, rrows)
    _autosize(ws_results)

    return _to_bytes(wb)


def rankings_xlsx(rankings: Iterable, sport: str | None = None,
                  weights: dict | None = None) -> BytesIO:
    """Rankings workbook with optional ``Weights`` sheet."""
    wb = Workbook()
    ws_rank = wb.active
    ws_rank.title = "Rankings"
    title = f"Athlete Rankings — {sport.upper() if sport else 'All Sports'}"
    ws_rank["A1"] = title
    ws_rank["A1"].font = TITLE_FONT
    ws_rank.merge_cells("A1:D1")
    ws_rank["A2"] = "Pro Sports Talents | Performance ranking"
    ws_rank["A2"].font = Font(italic=True, color="FF6B7280")
    ws_rank.merge_cells("A2:D2")

    rankings = list(rankings)
    rows = []
    for i, r in enumerate(rankings, start=1):
        if isinstance(r, dict):
            rank = r.get("rank", i)
            name = r.get("name", "")
            team = r.get("team", "") or ""
            score = r.get("score")
        else:
            rank = i
            name = getattr(r, "name", "")
            team = getattr(r, "team", "") or ""
            score = getattr(r, "score", None)
        rows.append([rank, name, team, score])
    _write_table(ws_rank, 4, ["Rank", "Athlete", "Team", "Score"], rows)
    _autosize(ws_rank)

    if weights:
        ws_w = wb.create_sheet("Weights")
        wrows = [[k, v] for k, v in weights.items()]
        _write_table(ws_w, 1, ["Metric", "Weight"], wrows)
        _autosize(ws_w)

    return _to_bytes(wb)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _to_bytes(wb: Workbook) -> BytesIO:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _team_label(team) -> str:
    if team is None:
        return ""
    return (
        getattr(team, "full_name", None)
        or getattr(team, "name", None)
        or getattr(team, "abbreviation", None)
        or ""
    )


def _recent_games_for(athlete, limit: int = 20):
    try:
        from app.models import NBAGame, NBATeam, NHLGame, NHLTeam
    except Exception:
        return []
    if not (getattr(athlete, "current_team", None) and getattr(athlete, "primary_sport", None)):
        return []
    code = athlete.primary_sport.code
    try:
        if code == "NBA":
            team = NBATeam.query.filter(
                (NBATeam.name.ilike(athlete.current_team))
                | (NBATeam.full_name.ilike(athlete.current_team))
                | (NBATeam.abbreviation.ilike(athlete.current_team))
            ).first()
            if team:
                return (
                    NBAGame.query
                    .filter((NBAGame.home_team_id == team.team_id)
                            | (NBAGame.visitor_team_id == team.team_id))
                    .order_by(NBAGame.date.desc())
                    .limit(limit)
                    .all()
                )
        elif code == "NHL":
            team = NHLTeam.query.filter(
                (NHLTeam.name.ilike(athlete.current_team))
                | (NHLTeam.location.ilike(athlete.current_team))
                | (NHLTeam.abbreviation.ilike(athlete.current_team))
            ).first()
            if team:
                return (
                    NHLGame.query
                    .filter((NHLGame.home_team_id == team.team_id)
                            | (NHLGame.visitor_team_id == team.team_id))
                    .order_by(NHLGame.date.desc())
                    .limit(limit)
                    .all()
                )
    except Exception:
        return []
    return []
